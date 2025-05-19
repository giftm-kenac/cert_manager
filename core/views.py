import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import HttpResponseForbidden, HttpResponseNotFound, JsonResponse, HttpResponse, HttpResponseRedirect, \
    Http404
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.db import transaction, IntegrityError
from django.urls import reverse, reverse_lazy
import uuid
import secrets

from django.conf import settings
User = settings.AUTH_USER_MODEL

from users.decorators import client_required, employee_required
from users.models import CustomUser, ClientProfile
from django.contrib.auth.hashers import make_password
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, UpdateView

from .models import (
    Certificate, CertificationType, TrainingCourse, Schedule, Skill,
    Event, EventQuestion, EventRegistration, EventRegistrationAnswer
)
from .forms import (
    CertificationTypeForm, TrainingCourseForm, IssueCertificateForm,
    ScheduleForm, UpdateScheduleStatusForm, VerifyByIdForm, SkillForm,
    BulkClientUploadForm, BulkIssueCertificateForm,
    EventForm, EventQuestionFormSet, PublicEventRegistrationForm,
    EventAttendeeSelectionForm, EventNotificationForm
)
from .email_utils import (
    send_certificate_issued_email_html,
    generate_certificate_image_with_details,
    send_employee_welcome_email_html,
    render_certificate_to_pdf,
    send_event_registration_confirmation_email,
    send_new_user_credentials_event_email,
    send_event_notification_to_registrants,
    # send_client_welcome_email_html, # Ensure this exists in email_utils.py if used below
    # send_admin_event_cancellation_notice_email, # Placeholder
    # send_user_registration_cancellation_confirmation_email # Placeholder
)


@client_required
def client_dashboard_view(request):
    total_certificates = Certificate.objects.filter(client=request.user).count()
    active_schedules_count = Schedule.objects.filter(
        client=request.user,
        status__in=['REQUESTED', 'SCHEDULED']
    ).count()
    recent_certificates = Certificate.objects.filter(client=request.user).order_by('-issue_date')[:5]
    active_event_registrations = EventRegistration.objects.filter(
        user=request.user
    ).exclude(status__in=['CANCELLED_BY_USER', 'CANCELLED_BY_ADMIN', 'NOT_ATTENDED']).select_related('event').order_by('event__date')[:5]


    context = {
        'total_certificates': total_certificates,
        'active_schedules_count': active_schedules_count,
        'recent_certificates': recent_certificates,
        'active_event_registrations': active_event_registrations,
    }
    return render(request, 'core/client_dashboard.html', context)


@client_required
def my_certificates_view(request):
    certificates = Certificate.objects.filter(client=request.user).select_related('certification_type').order_by('-issue_date')
    context = {'certificates': certificates}
    return render(request, 'core/my_certificates.html', context)


@client_required
def my_schedule_view(request):
    schedules = Schedule.objects.filter(client=request.user).select_related('course').order_by('-event_datetime', '-created_at')
    context = {'schedules': schedules}
    return render(request, 'core/my_schedule.html', context)


@client_required
def available_courses_view(request):
    courses = TrainingCourse.objects.filter(is_active=True).select_related('certification_type').order_by('start_date', 'name')
    context = {'courses': courses}
    return render(request, 'core/available_courses.html', context)


@client_required
def course_detail_view(request, course_id):
    course = get_object_or_404(TrainingCourse, pk=course_id, is_active=True)
    existing_schedule = Schedule.objects.filter(client=request.user, course=course).exclude(status__icontains='CANCELLED').first()
    schedule_form = ScheduleForm()

    if request.method == 'POST':
        if 'cancel_schedule' in request.POST and existing_schedule:
            if existing_schedule.status in ['REQUESTED', 'SCHEDULED']:
                existing_schedule.status = 'CANCELLED_BY_USER'
                existing_schedule.updated_at = timezone.now()
                existing_schedule.save(update_fields=['status', 'updated_at'])
                messages.info(request, f"Your schedule request for {course.name} has been cancelled.")
            else:
                messages.warning(request, "This schedule cannot be cancelled at its current status.")
            return redirect('core:course_detail', course_id=course.id)
        else:
            schedule_form = ScheduleForm(request.POST)
            if schedule_form.is_valid():
                if existing_schedule:
                    messages.warning(request, f"You already have an active schedule for {course.name}.")
                else:
                    event_dt = schedule_form.cleaned_data['event_datetime']
                    notes = schedule_form.cleaned_data.get('notes')
                    if event_dt < timezone.now():
                        schedule_form.add_error('event_datetime', "Selected date and time must be in the future.")
                    else:
                        Schedule.objects.create(client=request.user, course=course, event_datetime=event_dt, notes=notes, status='REQUESTED')
                        messages.success(request, f"Your request for {course.name} on {event_dt.strftime('%Y-%m-%d %H:%M')} has been submitted.")
                        return redirect('core:course_detail', course_id=course.id)
    context = {'course': course, 'existing_schedule': existing_schedule, 'schedule_form': schedule_form}
    return render(request, 'core/course_detail.html', context)


def verify_certificate_view(request, certificate_id):
    try:
        certificate_uuid = uuid.UUID(str(certificate_id))
        certificate = Certificate.objects.select_related('client__clientprofile', 'certification_type').get(id=certificate_uuid)
    except (Certificate.DoesNotExist, ValueError):
        return render(request, 'core/verify_certificate_not_found.html', status=404)
    context = {'certificate': certificate, 'is_valid': not certificate.is_expired()}
    return render(request, 'core/verify_certificate.html', context)


def verify_certificate_by_id_input_view(request):
    certificate = None
    not_found = False
    image_missing = False
    form = VerifyByIdForm()
    if request.method == 'POST':
        form = VerifyByIdForm(request.POST)
        if form.is_valid():
            cert_id_input = form.cleaned_data['certificate_id']
            try:
                certificate = Certificate.objects.select_related('client__clientprofile', 'certification_type').get(id=cert_id_input)
                if not certificate.generated_certificate_image:
                    image_missing = True
            except Certificate.DoesNotExist:
                not_found = True
            except ValueError:
                messages.error(request, "Invalid Certificate ID format.")
                not_found = True
    context = {'form': form, 'certificate': certificate, 'not_found': not_found, 'image_missing': image_missing}
    return render(request, 'core/verify_by_id_form.html', context)


@employee_required
def employee_dashboard_view(request):
    total_clients = CustomUser.objects.filter(is_employee=False).count()
    total_certificates_issued = Certificate.objects.count()
    active_courses = TrainingCourse.objects.filter(is_active=True).count()
    upcoming_events = Event.objects.filter(date__gte=timezone.now().date(), is_active=True).order_by('date', 'time').count()
    context = {
        'total_clients': total_clients,
        'total_certificates_issued': total_certificates_issued,
        'active_courses': active_courses,
        'upcoming_events': upcoming_events,
    }
    return render(request, 'core/employee_dashboard.html', context)


@employee_required
def manage_certification_types_view(request):
    types = CertificationType.objects.all().order_by('name')
    context = {'certification_types': types}
    return render(request, 'core/manage_certification_types.html', context)


@employee_required
def add_certification_type_view(request):
    if request.method == 'POST':
        form = CertificationTypeForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Certification Type added successfully.")
            return redirect('core:manage_certification_types')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = CertificationTypeForm()
    context = {'form': form}
    return render(request, 'core/add_certification_type.html', context)


@employee_required
def edit_certification_type_view(request, certification_type_id):
    cert_type = get_object_or_404(CertificationType, pk=certification_type_id)
    if request.method == 'POST':
        form = CertificationTypeForm(request.POST, request.FILES, instance=cert_type)
        if form.is_valid():
            form.save()
            messages.success(request, f"Certification Type '{cert_type.name}' updated successfully.")
            return redirect('core:manage_certification_types')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = CertificationTypeForm(instance=cert_type)
    context = {'form': form, 'cert_type': cert_type}
    return render(request, 'core/edit_certification_type.html', context)


@employee_required
def manage_courses_view(request):
    if request.method == 'POST':
        form = TrainingCourseForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Training Course added successfully.")
            return redirect('core:manage_courses')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = TrainingCourseForm()
    courses = TrainingCourse.objects.select_related('certification_type').all().order_by('name')
    context = {'form': form, 'courses': courses}
    return render(request, 'core/manage_courses.html', context)


@employee_required
@transaction.atomic
def issue_certificate_view(request):
    if request.method == 'POST':
        form = IssueCertificateForm(request.POST)
        if form.is_valid():
            client = form.cleaned_data['client']
            certification_type = form.cleaned_data['certification_type']
            if Certificate.objects.filter(client=client, certification_type=certification_type).exists():
                messages.warning(request, f"{client.get_full_name()} already has the certificate '{certification_type.name}'.")
            else:
                try:
                    certificate = Certificate(client=client, certification_type=certification_type, issue_date=form.cleaned_data.get('issue_date') or timezone.now().date())
                    certificate.save()
                    send_certificate_issued_email_html(to_email=client.email, fullname=client.get_full_name(), certificate_name=certification_type.name, certificate_instance=certificate)
                    messages.success(request, f"Certificate '{certification_type.name}' issued successfully to {client.get_full_name()}.")
                    return redirect('core:issue_certificate')
                except Exception as e:
                    messages.error(request, f"Failed to issue certificate: {e}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = IssueCertificateForm()
    context = {'form': form}
    return render(request, 'core/issue_certificate.html', context)


@employee_required
def employee_course_detail_view(request, course_id):
    course = get_object_or_404(TrainingCourse, pk=course_id)
    schedules = Schedule.objects.filter(course=course).select_related('client__clientprofile').order_by('-created_at')
    if request.method == 'POST' and 'update_status' in request.POST:
        status_form = UpdateScheduleStatusForm(request.POST)
        if status_form.is_valid():
            schedule_id = status_form.cleaned_data['schedule_id']
            new_status = status_form.cleaned_data['status']
            try:
                schedule_to_update = Schedule.objects.get(id=schedule_id, course=course)
                schedule_to_update.status = new_status
                schedule_to_update.updated_at = timezone.now()
                schedule_to_update.save(update_fields=['status', 'updated_at'])
                messages.success(request, f"Status updated for {schedule_to_update.client.get_full_name()}.")
                return redirect('core:employee_course_detail', course_id=course.id)
            except Schedule.DoesNotExist:
                messages.error(request, "Schedule not found.")
            except Exception as e:
                messages.error(request, f"Error updating status: {e}")
        else:
            messages.error(request, "Invalid status update request.")
    schedule_forms = []
    for schedule in schedules:
        form_instance = UpdateScheduleStatusForm(initial={'status': schedule.status, 'schedule_id': schedule.id})
        schedule_forms.append({'schedule': schedule, 'form': form_instance})
    context = {'course': course, 'schedule_forms': schedule_forms}
    return render(request, 'core/employee_course_detail.html', context)


@employee_required
def edit_course_view(request, course_id):
    course = get_object_or_404(TrainingCourse, pk=course_id)
    if request.method == 'POST':
        form = TrainingCourseForm(request.POST, request.FILES, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, f"Course '{course.name}' updated successfully.")
            return redirect('core:manage_courses')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = TrainingCourseForm(instance=course)
    context = {'form': form, 'course': course}
    return render(request, 'core/edit_course.html', context)


@employee_required
def manage_clients_view(request):
    upload_form = BulkClientUploadForm()
    upload_errors = request.session.pop('bulk_upload_errors', None)
    if request.method == 'POST':
        upload_form = BulkClientUploadForm(request.POST, request.FILES)
        if upload_form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                header = [cell.value for cell in sheet[1]]
                required_headers = ['Email', 'FirstName', 'LastName']
                if not all(h in header for h in required_headers):
                    messages.error(request, f"Excel file missing required columns: {', '.join(required_headers)}")
                    return redirect('core:manage_clients')
                col_map = {name: i for i, name in enumerate(header)}
                created_count = 0
                skipped_count = 0
                error_list = []
                with transaction.atomic():
                    for row_idx in range(2, sheet.max_row + 1):
                        row_values = [cell.value for cell in sheet[row_idx]]
                        email = row_values[col_map['Email']]
                        first_name = row_values[col_map['FirstName']]
                        last_name = row_values[col_map['LastName']]
                        if not email or not first_name or not last_name:
                            error_list.append(f"Row {row_idx}: Missing required data.")
                            skipped_count += 1
                            continue
                        if CustomUser.objects.filter(email__iexact=email).exists():
                            error_list.append(f"Row {row_idx}: Email '{email}' already exists.")
                            skipped_count += 1
                            continue
                        try:
                            password = CustomUser.objects.make_random_password()
                            user = CustomUser.objects.create_user(email=email, first_name=str(first_name), last_name=str(last_name), password=password, phone_number=str(row_values[col_map.get('PhoneNumber')]) if 'PhoneNumber' in col_map and row_values[col_map.get('PhoneNumber')] else None, is_employee=False, is_verified=True) # Auto-verify for now
                            ClientProfile.objects.create(user=user, organization=str(row_values[col_map.get('Organization')]) if 'Organization' in col_map and row_values[col_map.get('Organization')] else None, address=str(row_values[col_map.get('Address')]) if 'Address' in col_map and row_values[col_map.get('Address')] else None, city=str(row_values[col_map.get('City')]) if 'City' in col_map and row_values[col_map.get('City')] else None, country=str(row_values[col_map.get('Country')]) if 'Country' in col_map and row_values[col_map.get('Country')] else None, date_of_birth=row_values[col_map.get('DOB')] if 'DOB' in col_map else None, gender=str(row_values[col_map.get('Gender')]) if 'Gender' in col_map and row_values[col_map.get('Gender')] else None)
                            send_employee_welcome_email_html(user.email, user.get_full_name(), password) # Using employee welcome for simplicity, adapt if needed
                            created_count += 1
                        except Exception as e:
                            error_list.append(f"Row {row_idx} (Email: {email}): Error - {e}")
                            skipped_count += 1
                if created_count > 0: messages.success(request, f"Successfully created {created_count} client accounts.")
                if skipped_count > 0: messages.warning(request, f"Skipped {skipped_count} rows.")
                if error_list:
                    request.session['bulk_upload_errors'] = error_list[:10]
                    messages.error(request, "Some rows could not be processed.")
                return redirect('core:manage_clients')
            except Exception as e:
                messages.error(request, f"Error processing Excel file: {e}")
        else:
            messages.error(request, "Invalid file submitted.")
    clients = CustomUser.objects.filter(is_employee=False).select_related('clientprofile').order_by('last_name', 'first_name')
    context = {'clients': clients, 'upload_form': upload_form, 'upload_errors': upload_errors}
    return render(request, 'core/manage_clients.html', context)


@employee_required
def issued_certificates_view(request):
    certificates = Certificate.objects.select_related('client__clientprofile', 'certification_type').all().order_by('-issue_date')
    context = {'certificates': certificates}
    return render(request, 'core/issued_certificates.html', context)


@employee_required
def client_detail_view(request, client_user_id):
    client_user = get_object_or_404(CustomUser, pk=client_user_id, is_employee=False)
    client_profile = ClientProfile.objects.filter(user=client_user).first()
    client_certificates = Certificate.objects.filter(client=client_user).select_related('certification_type').order_by('-issue_date')
    client_schedules = Schedule.objects.filter(client=client_user).select_related('course').order_by('-created_at')
    client_event_registrations = EventRegistration.objects.filter(user=client_user).select_related('event').order_by('-event__date')
    context = {'client_user': client_user, 'client_profile': client_profile, 'certificates': client_certificates, 'schedules': client_schedules, 'event_registrations': client_event_registrations}
    return render(request, 'core/client_detail.html', context)


@login_required
def download_certificate_image_view(request, certificate_id):
    certificate_uuid = uuid.UUID(str(certificate_id))
    certificate = get_object_or_404(Certificate, id=certificate_uuid)
    if not request.user.is_employee and certificate.client != request.user:
        messages.error(request, "You are not authorized to download this certificate.")
        return redirect('core:client_dashboard')
    try:
        generate_certificate_image_with_details(certificate)
        certificate.save(update_fields=['generated_certificate_image'])
        if certificate.generated_certificate_image and hasattr(certificate.generated_certificate_image, 'path'):
            with open(certificate.generated_certificate_image.path, 'rb') as f:
                image_buffer = f.read()
            response = HttpResponse(image_buffer, content_type='image/png')
            filename = f"Certificate_{certificate.certification_type.name.replace(' ', '_')}_{certificate.client.get_full_name().replace(' ', '_')}_{certificate.id}.png"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        else:
            raise ValueError("Generated image not found or path is missing.")
    except Exception as e:
        messages.error(request, f"Could not generate or serve certificate image: {e}")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', reverse('core:my_certificates')))


@employee_required
def manage_skills_view(request):
    if request.method == 'POST':
        form = SkillForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Skill added successfully.")
                return redirect('core:manage_skills')
            except IntegrityError:
                messages.error(request, f"Skill '{form.cleaned_data['name']}' already exists.")
            except Exception as e:
                messages.error(request, f"Failed to add skill: {e}")
        else:
            messages.error(request, "Please correct the error below.")
    else:
        form = SkillForm()
    skills = Skill.objects.all().order_by('name')
    context = {'form': form, 'skills': skills}
    return render(request, 'core/manage_skills.html', context)


@employee_required
def edit_skill_view(request, skill_id):
    skill = get_object_or_404(Skill, pk=skill_id)
    if request.method == 'POST':
        form = SkillForm(request.POST, instance=skill)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, f"Skill '{skill.name}' updated successfully.")
                return redirect('core:manage_skills')
            except IntegrityError:
                messages.error(request, f"A skill with the name '{form.cleaned_data['name']}' already exists.")
            except Exception as e:
                messages.error(request, f"Failed to update skill: {e}")
        else:
            messages.error(request, "Please correct the error below.")
    else:
        form = SkillForm(instance=skill)
    context = {'form': form, 'skill': skill}
    return render(request, 'core/edit_skill.html', context)


@employee_required
def delete_skill_view(request, skill_id):
    skill = get_object_or_404(Skill, pk=skill_id)
    if request.method == 'POST':
        try:
            skill_name = skill.name
            skill.delete()
            messages.success(request, f"Skill '{skill_name}' deleted successfully.")
        except Exception as e:
            messages.error(request, f"Could not delete skill '{skill.name}'. It might be in use. Error: {e}")
        return redirect('core:manage_skills')
    else:
        messages.warning(request, "Deletion must be done via POST request.")
        return redirect('core:manage_skills')


@login_required
def download_certificate_pdf_view(request, certificate_id):
    certificate_uuid = uuid.UUID(str(certificate_id))
    certificate = get_object_or_404(Certificate, pk=certificate_uuid)
    if not request.user.is_employee and certificate.client != request.user:
        messages.error(request, "You are not authorized to download this certificate PDF.")
        return redirect('core:client_dashboard')
    try:
        pdf_content = render_certificate_to_pdf(certificate)
        if pdf_content:
            response = HttpResponse(pdf_content, content_type='application/pdf')
            filename = f"Certificate_{certificate.certification_type.name.replace(' ', '_')}_{certificate.client.get_full_name().replace(' ', '_')}_{certificate.id}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        else:
            messages.error(request, "Could not generate certificate PDF.")
    except Exception as e:
        messages.error(request, f"Error generating PDF: {e}")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', reverse('core:my_certificates')))


@employee_required
def bulk_issue_certificates_view(request):
    if request.method == 'POST':
        form = BulkIssueCertificateForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['email_file']
            certification_type = form.cleaned_data['certification_type']
            issue_date = form.cleaned_data.get('issue_date') or timezone.now().date()
            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                header = [cell.value for cell in sheet[1]]
                if 'Email' not in header:
                    messages.error(request, "Excel file must contain a column named 'Email'.")
                    return redirect('core:bulk_issue_certificates')
                email_col_index = header.index('Email')
                issued_count = 0
                skipped_count = 0
                error_list = []
                with transaction.atomic():
                    for row_idx in range(2, sheet.max_row + 1):
                        row_values = [cell.value for cell in sheet[row_idx]]
                        email = row_values[email_col_index]
                        if not email:
                            error_list.append(f"Row {row_idx}: Missing email address.")
                            skipped_count += 1
                            continue
                        try:
                            client = CustomUser.objects.get(email__iexact=email, is_employee=False)
                            if not client.is_verified:
                                error_list.append(f"Row {row_idx}: Client '{email}' is not verified.")
                                skipped_count += 1
                                continue
                            if Certificate.objects.filter(client=client, certification_type=certification_type).exists():
                                error_list.append(f"Row {row_idx}: Client '{email}' already has this certificate.")
                                skipped_count += 1
                                continue
                            certificate = Certificate(client=client, certification_type=certification_type, issue_date=issue_date)
                            certificate.save()
                            send_certificate_issued_email_html(to_email=client.email, fullname=client.get_full_name(), certificate_name=certification_type.name, certificate_instance=certificate)
                            issued_count += 1
                        except CustomUser.DoesNotExist:
                            error_list.append(f"Row {row_idx}: Client with email '{email}' not found.")
                            skipped_count += 1
                        except Exception as e:
                            error_list.append(f"Row {row_idx} (Email: {email}): Error - {e}")
                            skipped_count += 1
                if issued_count > 0: messages.success(request, f"Successfully issued {issued_count} certificates.")
                if skipped_count > 0: messages.warning(request, f"Skipped {skipped_count} rows.")
                if error_list:
                    request.session['bulk_issue_errors'] = error_list[:10]
                    messages.error(request, "Some rows could not be processed.")
                return redirect('core:bulk_issue_certificates')
            except Exception as e:
                messages.error(request, f"Error processing Excel file: {e}")
        else:
            messages.error(request, "Invalid form submission.")
    else:
        form = BulkIssueCertificateForm()
    upload_errors = request.session.pop('bulk_issue_errors', None)
    context = {'form': form, 'upload_errors': upload_errors}
    return render(request, 'core/bulk_issue_certificates.html', context)


@method_decorator(employee_required, name='dispatch')
class EventListView(ListView):
    model = Event
    template_name = 'core/events/event_list_admin.html'
    context_object_name = 'events'
    paginate_by = 10
    def get_queryset(self):
        return Event.objects.all().order_by('-date', '-time')

@method_decorator(employee_required, name='dispatch')
class EventCreateView(CreateView):
    model = Event
    form_class = EventForm
    template_name = 'core/events/event_form.html'
    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['questions_formset'] = EventQuestionFormSet(self.request.POST, prefix='questions')
        else:
            data['questions_formset'] = EventQuestionFormSet(prefix='questions')
        data['form_title'] = "Create New Event"
        return data
    def form_valid(self, form):
        context = self.get_context_data()
        questions_formset = context['questions_formset']
        with transaction.atomic():
            form.instance.created_by = self.request.user
            self.object = form.save()
            if questions_formset.is_valid():
                questions_formset.instance = self.object
                questions_formset.save()
                messages.success(self.request, f"Event '{self.object.name}' created successfully.")
            else:
                messages.error(self.request, "Please correct the errors in the questions section.")
                return self.form_invalid(form)
        return super().form_valid(form)
    def get_success_url(self):
        return reverse('event_detail_admin', kwargs={'pk': self.object.pk})

@method_decorator(employee_required, name='dispatch')
class EventUpdateView(UpdateView):
    model = Event
    form_class = EventForm
    template_name = 'core/events/event_form.html'
    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['questions_formset'] = EventQuestionFormSet(self.request.POST, instance=self.object, prefix='questions')
        else:
            data['questions_formset'] = EventQuestionFormSet(instance=self.object, prefix='questions')
        data['form_title'] = f"Edit Event: {self.object.name}"
        return data
    def form_valid(self, form):
        context = self.get_context_data()
        questions_formset = context['questions_formset']
        with transaction.atomic():
            self.object = form.save()
            if questions_formset.is_valid():
                questions_formset.instance = self.object
                questions_formset.save()
                messages.success(self.request, f"Event '{self.object.name}' updated successfully.")
            else:
                messages.error(self.request, "Please correct the errors in the questions section.")
                return self.form_invalid(form)
        return super().form_valid(form)
    def get_success_url(self):
        return reverse('event_detail_admin', kwargs={'pk': self.object.pk})

@method_decorator(employee_required, name='dispatch')
class EventDetailAdminView(DetailView):
    model = Event
    template_name = 'core/events/event_detail_admin.html'
    context_object_name = 'event'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        event = self.get_object()
        context['registrations'] = EventRegistration.objects.filter(event=event).select_related('user__client_profile').order_by('-registration_date')
        context['questions'] = EventQuestion.objects.filter(event=event).prefetch_related('options').order_by('order')
        if event.certification_type:
            context['attendee_form'] = EventAttendeeSelectionForm(event_id=event.pk)
        context['notification_form'] = EventNotificationForm()
        return context

@method_decorator(employee_required, name='dispatch')
class AdminCancelEventView(View):
    def post(self, request, event_pk):
        event = get_object_or_404(Event, pk=event_pk)
        if not event.is_active:
            messages.warning(request, f"Event '{event.name}' is already inactive/cancelled.")
            return redirect('event_detail_admin', pk=event.pk)
        with transaction.atomic():
            event.is_active = False
            event.save(update_fields=['is_active'])
            registrations_to_notify = EventRegistration.objects.filter(event=event, status__in=['REGISTERED', 'CONFIRMED', 'WAITLISTED', 'ATTENDED']).select_related('user')
            updated_count = 0
            for reg in registrations_to_notify:
                reg.status = 'CANCELLED_BY_ADMIN'
                reg.save(update_fields=['status'])
                updated_count +=1
                # try:
                #     send_admin_event_cancellation_notice_email(reg.user.email, event) # Ensure this email util exists
                # except Exception as e:
                #     print(f"Failed to send cancellation notice to {reg.user.email} for event {event.name}: {e}")
        messages.success(request, f"Event '{event.name}' has been cancelled. {updated_count} registrations were updated.")
        return redirect('event_detail_admin', pk=event.pk)

class PublicEventRegisterView(View):
    template_name = 'core/events/public_event_registration.html'
    def get_event(self, event_slug):
        try:
            return Event.objects.prefetch_related('questions__options').get(slug=event_slug, is_active=True)
        except Event.DoesNotExist:
            raise Http404("Event not found or registration is closed.")
    def get(self, request, event_slug):
        event = self.get_event(event_slug)
        if event.registration_deadline and timezone.now() > event.registration_deadline:
            messages.error(request, f"Registration for '{event.name}' has closed.")
            return render(request, 'core/events/event_registration_closed.html', {'event': event})
        if request.user.is_authenticated:
            existing_registration = EventRegistration.objects.filter(event=event, user=request.user).first()
            if existing_registration and existing_registration.status not in ['CANCELLED_BY_USER', 'CANCELLED_BY_ADMIN']:
                messages.info(request, f"You are already registered for '{event.name}'. Status: {existing_registration.get_status_display()}")
        form = PublicEventRegistrationForm(event_questions=event.questions.all().order_by('order'))
        return render(request, self.template_name, {'event': event, 'form': form})
    @transaction.atomic
    def post(self, request, event_slug):
        event = self.get_event(event_slug)
        if event.registration_deadline and timezone.now() > event.registration_deadline:
            messages.error(request, f"Registration for '{event.name}' has closed.")
            return redirect('public_event_register', event_slug=event.slug)
        if event.max_attendees is not None:
            current_registrations_count = EventRegistration.objects.filter(event=event, status__in=['REGISTERED', 'CONFIRMED', 'ATTENDED']).count()
            if current_registrations_count >= event.max_attendees:
                messages.error(request, f"Sorry, registration for '{event.name}' is full.")
                return redirect('public_event_register', event_slug=event.slug)
        form = PublicEventRegistrationForm(request.POST, event_questions=event.questions.all().order_by('order'))
        if form.is_valid():
            user_email = form.cleaned_data['user_email']
            user = None
            new_user_created = False
            password = None
            if request.user.is_authenticated and request.user.email == user_email:
                user = request.user
            else:
                try:
                    user = User.objects.get(email=user_email)
                except User.DoesNotExist:
                    password = secrets.token_urlsafe(12)
                    user_data = {'email': user_email, 'password': password}
                    if hasattr(User, 'is_client'): user_data['is_client'] = True # Check User model definition
                    if hasattr(User, 'is_verified'): user_data['is_verified'] = True # Auto-verify for event registration
                    user = User.objects.create_user(**user_data)
                    if hasattr(user, 'is_client') and user.is_client:
                        ClientProfile.objects.get_or_create(user=user)
                    new_user_created = True
            if EventRegistration.objects.filter(event=event, user=user).exclude(status__in=['CANCELLED_BY_USER', 'CANCELLED_BY_ADMIN']).exists():
                messages.warning(request, f"You are already actively registered for {event.name}.")
                return redirect('public_event_register', event_slug=event.slug)
            registration = EventRegistration.objects.create(event=event, user=user, status='REGISTERED')
            for question in event.questions.all():
                field_name = f'question_{question.id}'
                answer_data = form.cleaned_data.get(field_name)
                if answer_data is not None:
                    reg_answer = EventRegistrationAnswer(registration=registration, question=question)
                    if question.field_type == 'checkbox':
                        reg_answer.save()
                        selected_option_ids = [int(opt_id) for opt_id in answer_data]
                        reg_answer.selected_options.set(selected_option_ids)
                    else:
                        reg_answer.answer_text = str(answer_data)
                        reg_answer.save()
            if new_user_created and password:
                send_new_user_credentials_event_email(user, password, event)
            send_event_registration_confirmation_email(user.email, event)
            messages.success(request, f"Successfully registered for {event.name}! A confirmation email has been sent.")
            return redirect('event_registration_success', event_slug=event.slug)
        else:
            return render(request, self.template_name, {'event': event, 'form': form})

class EventRegistrationSuccessView(View):
    template_name = 'core/events/event_registration_success.html'
    def get(self, request, event_slug):
        event = get_object_or_404(Event, slug=event_slug)
        return render(request, self.template_name, {'event': event})

@method_decorator(login_required, name='dispatch')
class UserCancelEventRegistrationView(View):
    def post(self, request, registration_pk):
        registration = get_object_or_404(EventRegistration, pk=registration_pk, user=request.user)
        event = registration.event
        if registration.status in ['CANCELLED_BY_USER', 'CANCELLED_BY_ADMIN']:
            messages.warning(request, f"Your registration for '{event.name}' is already cancelled.")
        elif event.registration_deadline and timezone.now() > event.registration_deadline :
            messages.error(request, f"The deadline to cancel registration for '{event.name}' has passed.")
        elif not event.is_active:
            messages.error(request, f"Cannot cancel registration as event '{event.name}' is no longer active or has been cancelled by the administrator.")
        else:
            registration.status = 'CANCELLED_BY_USER'
            registration.save(update_fields=['status'])
            # try:
            #     send_user_registration_cancellation_confirmation_email(request.user.email, event) # Ensure this email util exists
            # except Exception as e:
            #     print(f"Failed to send user cancellation confirmation to {request.user.email} for event {event.name}: {e}")
            messages.success(request, f"Your registration for '{event.name}' has been cancelled.")
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', reverse('core:client_dashboard')))

@employee_required
def manage_event_registrations(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    registrations = EventRegistration.objects.filter(event=event).select_related('user__clientprofile').order_by('user__email')
    if request.method == 'POST':
        for reg_id_key, status_value in request.POST.items():
            if reg_id_key.startswith('status_'):
                reg_pk = reg_id_key.split('_')[1]
                try:
                    registration_to_update = EventRegistration.objects.get(pk=reg_pk, event=event)
                    if status_value in [choice[0] for choice in EventRegistration.STATUS_CHOICES]:
                        registration_to_update.status = status_value
                        if status_value == 'ATTENDED': registration_to_update.attended = True
                        elif status_value == 'NOT_ATTENDED': registration_to_update.attended = False
                        registration_to_update.save(update_fields=['status', 'attended'])
                except EventRegistration.DoesNotExist:
                    messages.error(request, f"Registration ID {reg_pk} not found for this event.")
                except Exception as e:
                    messages.error(request, f"Error updating registration {reg_pk}: {e}")
        messages.success(request, "Registration statuses updated.")
        return redirect('manage_event_registrations', event_pk=event.pk)
    context = {'event': event, 'registrations': registrations, 'status_choices': EventRegistration.STATUS_CHOICES}
    return render(request, 'core/events/manage_event_registrations.html', context)

@employee_required
@transaction.atomic
def issue_event_certificates(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    if not event.certification_type:
        messages.error(request, "This event does not have a certification type assigned. Certificates cannot be issued.")
        return redirect('core:event_detail_admin', pk=event.pk)
    eligible_registrations_qs = EventRegistration.objects.filter(event=event, status='ATTENDED').select_related('user', 'user__clientprofile')
    if request.method == 'POST':
        form = EventAttendeeSelectionForm(request.POST, event_id=event.pk)
        form.fields['attendees'].queryset = eligible_registrations_qs
        if form.is_valid():
            selected_registrations = form.cleaned_data['attendees']
            issued_count = 0
            failed_count = 0
            for registration in selected_registrations:
                try:
                    if Certificate.objects.filter(client=registration.user, certification_type=event.certification_type).exists():
                        messages.warning(request, f"Certificate already exists for {registration.user.email} for '{event.certification_type.name}'.")
                        continue
                    new_cert = Certificate(client=registration.user, certification_type=event.certification_type, issued_date=timezone.now().date())
                    new_cert.save()
                    send_certificate_issued_email_html(to_email=registration.user.email, fullname=registration.user.get_full_name() if hasattr(registration.user, 'get_full_name') else registration.user.email, certificate_name=event.certification_type.name, certificate_instance=new_cert)
                    issued_count += 1
                except Exception as e:
                    failed_count += 1
                    messages.error(request, f"Failed to issue certificate for {registration.user.email}: {e}")
            if issued_count > 0: messages.success(request, f"Successfully processed {issued_count} certificates.")
            if failed_count > 0: messages.warning(request, f"Failed to issue {failed_count} certificates.")
            return redirect('event_detail_admin', pk=event.pk)
    else:
        form = EventAttendeeSelectionForm(event_id=event.pk)
        form.fields['attendees'].queryset = eligible_registrations_qs
    context = {'event': event, 'form': form, 'eligible_registrations_count': eligible_registrations_qs.count()}
    return render(request, 'core/events/issue_event_certificates.html', context)

@employee_required
def send_event_notification_view(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    if request.method == 'POST':
        form = EventNotificationForm(request.POST)
        if form.is_valid():
            subject = form.cleaned_data['subject']
            message_body_html = form.cleaned_data['message']
            registrations = EventRegistration.objects.filter(event=event, status__in=['REGISTERED', 'CONFIRMED', 'ATTENDED', 'WAITLISTED'])
            recipient_list = [reg.user.email for reg in registrations if reg.user.email]
            if recipient_list:
                send_event_notification_to_registrants(event, subject, message_body_html, recipient_list)
                messages.success(request, f"Notification sent to {len(recipient_list)} registrants for {event.name}.")
            else:
                messages.warning(request, "No active registrants found to send notifications to.")
            return redirect('event_detail_admin', pk=event.pk)
    else:
        form = EventNotificationForm()
    context = {'event': event, 'form': form}
    return render(request, 'core/events/send_event_notification.html', context)